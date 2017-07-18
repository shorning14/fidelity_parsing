import csv, sys, collections, string, openpyxl, yagmail, os, multiprocessing as mp
from openpyxl.utils.cell import get_column_letter
from decimal import *
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference, 
    BarChart
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList 



#Sean Horning
#6/28/17
#Program for generating raw data accounting excel sheet from fidelity investments csv files

#Define the static filepaths and reference dictionaries
dest_file = r'C:\Program Files\Notepad++\fdestination.csv'
read_file = r'C:\Program Files\Notepad++\position.csv'
client_file = r'C:\Program Files\Notepad++\client_export.csv'
account_file = r'C:\Program Files\Notepad++\account.csv'
vdict = collections.defaultdict(int)
cdict = collections.defaultdict(str)
ndict = collections.defaultdict(str)



#Division helper that returns 0 when the denominator is 0 and 
#truncates the result to two decimal places.  
def safe_divide(num, denom):
	if denom==0:
		return 0
	else:
		return round(num/denom,2)

#Constructs a default dictionary mapping account numbers to their total market values
def value_mapping_helper():
	with open(read_file, 'rU') as z:
		reader3 = csv.reader(z)
		for row in reader3:
			vdict[row[0]]+= Decimal((row[11]))
			round(vdict[row[0]],2)



#Helper function that maps account numbers to client name for easy lookup
def client_mapping_helper():
	with open(client_file,'rU') as x:
		reader = csv.reader(x)
		for row in reader:
			if row[0] == 'C':
				current_client = string.replace(row[1],',','.')
			else:
				cdict[row[1]] = current_client





#Map account numbers to primary account owner: last name- first name - M.I
def name_mapping_helper():
	with open(account_file, 'rU') as y:
		reader = csv.reader(y)
		for row in reader:
			ndict[row[1]] = string.replace(string.replace(row[4],',','.'),'..','.') +'.' ' ' + row[2] +' '+ row[3]

#This function writes all of the records from the position csv and writes them in the 
#raw data format to a single csv
def transcribe_raw_data():
	#Open source and destination files 
	with open(read_file, 'rU') as f, open(dest_file, 'w+') as d:
		#Initialize csv reader
		reader = csv.reader(f)
		try:
			#Write all headers to destination file 
			d.write('Account#' + ',' + 'Custom Short Name' + ',' + 'FBSI Short Name' + ',' + 'Primary Account Owner' + ',' + 'Registration'+ ','+'Account Source'
			+ ',' + 'Account Type'+ ','+'Security Description'+ ','+'Symbol'+ ','+'Quantity'+ ','+'Closing Price'+ ','+'Market Value'+ ','+'Currency'+ ','+'Closing Price'
			+','+'Market Value'+ ','+'Intl Currency'+ ','+'% of Total'+ ','+'Cost Basis'+ ','+'Currency'+ ','+'Unrealized Gain/Loss'+ ','+'Currency' + ','+ 'Client' + '\n')
			#Fill in column values for each record
			for row in reader:
				d.write(row[0] + ',' + row[1] + ',' + string.replace(row[2],',','.') + ',' + ndict[row[0]] + ',' + row[4] + ',' + row[5] + ',' + row[6] + ',' + row[7] + ',' + row[8] + ','
				+str(round(float(row[9]),2)) + ',' + str(round(float(row[10]),2)) + ',' + str(round(float(row[11]),2)) + ',' + 'USD' + ',' + '' + ',' + '' + ',' + ' ' + ',' + str(safe_divide(Decimal(row[11]),vdict[row[0]])) + ',' 
				+row[12] + ',' + 'USD' + ',' + row[13] + ',' + 'USD' +','+ str(cdict[row[0]]) + '\n')
			f.close()
		#Error handling 
		except csv.Error as e:
			sys.exit('file %s, line %d: %s' % (read_file, reader.line_num, e))
	print('Raw data transcription completed')
	print('----------------------------------------')

#transcribe_client_reports takes the file creates by transcribe_raw_data and splits it into 
#many different csv files, so that each csv contains only account records relevant to a specific 
#client.
def transcribe_client_reports():
	for client in set(cdict.values()):
		report_path = r'C:\Program Files\Notepad++\reports' + '\\' + string.replace(client,'/','-') + '_report.csv'
		with open(report_path, 'wb') as d, open(dest_file, 'rb') as r:
			reader = csv.reader(r)
			writer = csv.writer(d)
			d.write('Account#' + ',' + 'Custom Short Name' + ',' + 'FBSI Short Name' + ',' + 'Primary Account Owner' + ',' + 'Registration'+ ','+'Account Source'
			+ ',' + 'Account Type'+ ','+'Security Description'+ ','+'Symbol'+ ','+'Quantity'+ ','+'Closing Price'+ ','+'Market Value'+ ','+'Currency'+ ','+'Closing Price'
			+','+'Market Value'+ ','+'Intl Currency'+ ','+'% of Total'+ ','+'Cost Basis'+ ','+'Currency'+ ','+'Unrealized Gain/Loss'+ ','+'Currency' + ','+ 'Client' + '\n')
			for row in reader:
				if row[21] == client:
					writer.writerow(row)
		d.close()
	print('Client report transcription completed')
	print('----------------------------------------')

#Workbook transcription, this function loads a workbook template, and fills the raw data worksheet 
#with all of the raw data associated with the client input
def transcribe_client_data_to_workbooks(client):
	report_path = r'C:\Program Files\Notepad++\reports' + '\\' + string.replace(client,'/','-') + '_report.csv'
	wb = openpyxl.load_workbook('C:\Program Files\Notepad++\Bucket-Asset Allocation Model.xlsm', read_only = False, keep_vba = True)
	wb.active = 5
	ws = wb.active
	with open(report_path, 'rU') as f:
		reader = csv.reader(f)
		for row_index, row in enumerate(reader):
			for column_index, cell in enumerate(row):
				column_letter = get_column_letter((column_index+1))
				if column_letter == 'A':
					if row_index > 0:
						ws[column_letter+str(row_index+1)] = long(string.replace(string.replace(cell,'Z','9'),'X','0'))
				else:
					ws[column_letter+str(row_index+1)] = cell
		#Proposed allocation pie chart creation
		#Close the file 
		f.close()

	#Proposed Allocation pie chart creation 
	wb.active = 1
	ws = wb.active
	pie1 = PieChart()
	labels1 = Reference(ws, min_col = 14, min_row = 19, max_row = 21)
	data1 = Reference(ws, min_col = 15, min_row = 18, max_row = 21)
	pie1.add_data(data1, titles_from_data = True)
	pie1.set_categories(labels1)
	pie1.title = 'Proposed Allocation'
	pie1.height = 14
	pie1.width = 18
	ws.add_chart(pie1, "F20" )
	pie1.dataLabels = DataLabelList()
	pie1.dataLabels.showPercent = True
	#Current allocation pie chart creation
	pie2 = PieChart()
	labels2 = Reference(ws, min_col = 14, min_row = 29, max_row = 31)
	data2 = Reference(ws, min_col = 15, min_row = 28, max_row = 31 )
	pie2.add_data(data2, titles_from_data = True)
	pie2.set_categories(labels2)
	pie2.title = 'Current Allocation'
	pie2.height = 14
	pie2.width = 18
	ws.add_chart(pie2, 'B20')
	pie2.dataLabels = DataLabelList()
	pie2.dataLabels.showPercent = True

	#Bar chart creation
	#Change to allocations worksheet
	wb.active = 3
	ws = wb.active
	#Short term bucket bar graph creation
	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = 12
	chart1.title = "Bucket I \n Short - Term"
	data = Reference(ws, min_col=14, min_row=61, max_row = 62, max_col=15)
	cats = Reference(ws, min_col=13, min_row=62)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = 4
	ws.add_chart(chart1, "B25")
	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True
	#Intermediate term bucket bar graph creation
	chart2 = BarChart()
	chart2.type = "col"
	chart2.style = 10
	chart2.title = "Bucket II \n Intermediate - Term"
	data2 = Reference(ws, min_col = 14, min_row = 63, max_row = 64, max_col = 15 )
	cats2 = Reference(ws, min_col = 13, min_row = 64)
	chart2.add_data(data2, titles_from_data = True)
	chart2.set_categories(cats2)
	chart2.shape = 4
	ws.add_chart(chart2, "I25" )
	chart2.dataLabels = DataLabelList()
	chart2.dataLabels.showVal = True
	#Long term bucket bar graph creation
	chart3 = BarChart()
	chart3.type = "col"
	chart3.style = 13
	chart3.title = "Bucket III \n Long - Term"
	data3 = Reference(ws, min_col = 14, min_row = 65, max_row = 66, max_col = 15 )
	cats3 = Reference(ws, min_col = 13, min_row = 66)
	chart3.add_data(data3, titles_from_data = True)
	chart3.set_categories(cats3)
	chart3.shape = 4
	ws.add_chart(chart3, "P25" )
	chart3.dataLabels = DataLabelList()
	chart3.dataLabels.showVal = True
	#0 portfolio bar graph creation
	chart4 = BarChart()
	chart4.type = "col"
	chart4.style = 10
	chart4.height = 10
	chart4.width = 20
	chart4.title = "0 Portfolio"
	data4 = Reference(ws, min_col = 14, min_row = 57, max_row = 58, max_col = 15)
	cats4 = Reference(ws, min_col = 14, min_row = 57, max_col = 15)
	chart4.add_data(data4, titles_from_data = True)
	chart4.set_categories(cats4)
	chart4.shape = 4
	ws.add_chart(chart4, "B3" )
	chart4.dataLabels = DataLabelList()
	chart4.dataLabels.showVal = True
	#Allocation Comparison Bar graph creation
	chart5 = BarChart()
	chart5.type = "col"
	chart5.style = 10
	chart5.title =  "Allocation Comparison"
	chart5.height = 10
	chart5.width = 20
	data5 = Reference(ws, min_col = 10, min_row = 62, max_row = 64, max_col = 12)
	cats5 = Reference(ws, min_col = 9, min_row = 63, max_row = 64)
	chart5.add_data(data5, titles_from_data = True)
	chart5.set_categories(cats5)
	chart5.shape = 4
	ws.add_chart(chart5, "M3" )
	chart5.dataLabels = DataLabelList()
	chart5.dataLabels.showVal = True
	print(len(set(cdict.values())))
	#Attempt to save, move on if the file is open for reading.  
	try:
		wb.save('C:\\Program Files\\Notepad++\\workbooks\\' + string.replace(client,'/','-') + '_workbook.xlsm')
		print(string.replace(client,'/','-')+' workbook completed')
	except IOError:
		print("Workbook already open, cannot overwrite, moving on.")
	#Print that the workbook has succesffuly been updated


for filename in os.listdir('C:\Program Files\Notepad++'):
	if '-(' in filename:
		os.remove(filename)

#If the current process reading this is the main process, calls the first 5 functions then 
#uses the specified number of process to map all clients to the transcribe_client_data_to_workbooks 
#function before terminating and printing a success.  
if __name__ == '__main__':
	value_mapping_helper()
	client_mapping_helper()
	name_mapping_helper()
	transcribe_raw_data()
	transcribe_client_reports()
	pool = mp.Pool(processes = 3)
	pool.map(transcribe_client_data_to_workbooks, set(cdict.values()))
	pool.terminate()
	yagmail.SMTP('jweaver@scandh.com"').send(["ghorning@scandh.com","jweaver@scandh.com","athompson@scandh.com","rpatalon@scandh.com",
											  "lprincipio@scandh.com"],"AUTOMATED: Fidelity Workbooks","The updated fidelity workbooks have completed generation.")
	print ("The Program has terminated successfully")
